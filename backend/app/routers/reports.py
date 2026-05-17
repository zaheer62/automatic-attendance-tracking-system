from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import cast, Date, func
from app.core.database import get_db
from app.models.attendance import AttendanceRecord, AttendanceSession
from app.models.user import User
from app.models.subject import Subject
import io
import openpyxl
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

router = APIRouter(prefix="/reports", tags=["Reports"])


# ── Shared helper ─────────────────────────────────────────────────────────────

def get_attendance_data(db: Session):
    students = db.query(User).filter(User.role == "student").all()
    data = []
    for student in students:
        total = db.query(AttendanceRecord).filter(
            AttendanceRecord.student_id == student.id
        ).count()
        present = db.query(AttendanceRecord).filter(
            AttendanceRecord.student_id == student.id,
            AttendanceRecord.is_present == True
        ).count()
        percentage = round((present / total * 100), 1) if total > 0 else 0
        data.append({
            "name": student.full_name,
            "email": student.email,
            "total": total,
            "present": present,
            "absent": total - present,
            "percentage": percentage
        })
    return data


# ── Daily report ──────────────────────────────────────────────────────────────

@router.get("/daily")
def daily_report(date: str, db: Session = Depends(get_db)):
    """
    Returns per-student attendance for a specific date (YYYY-MM-DD).
    """
    from datetime import datetime
    try:
        target_date = datetime.strptime(date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")

    sessions = db.query(AttendanceSession).filter(
        cast(AttendanceSession.session_date, Date) == target_date
    ).all()
    session_ids = [s.id for s in sessions]

    students = db.query(User).filter(User.role == "student").all()
    result = []
    for student in students:
        if session_ids:
            record = db.query(AttendanceRecord).filter(
                AttendanceRecord.student_id == student.id,
                AttendanceRecord.session_id.in_(session_ids)
            ).first()
        else:
            record = None

        result.append({
            "student_id": student.id,
            "student_name": student.full_name,
            "date": date,
            "status": ("present" if record and record.is_present
                       else "absent" if record
                       else "not_marked"),
            "method": record.method if record else None,
        })
    return result


# ── Monthly report ────────────────────────────────────────────────────────────

@router.get("/monthly")
def monthly_report(month: str, db: Session = Depends(get_db)):
    """
    Returns per-student totals for a month (YYYY-MM).
    """
    from datetime import datetime
    try:
        year, mon = map(int, month.split("-"))
    except (ValueError, AttributeError):
        raise HTTPException(status_code=400, detail="month must be YYYY-MM")

    # Sessions in that month
    sessions = db.query(AttendanceSession).filter(
        func.extract("year", AttendanceSession.session_date) == year,
        func.extract("month", AttendanceSession.session_date) == mon
    ).all()
    session_ids = [s.id for s in sessions]

    students = db.query(User).filter(User.role == "student").all()
    result = []
    for student in students:
        if session_ids:
            total = db.query(AttendanceRecord).filter(
                AttendanceRecord.student_id == student.id,
                AttendanceRecord.session_id.in_(session_ids)
            ).count()
            present = db.query(AttendanceRecord).filter(
                AttendanceRecord.student_id == student.id,
                AttendanceRecord.session_id.in_(session_ids),
                AttendanceRecord.is_present == True
            ).count()
        else:
            total = present = 0

        percentage = round(present / total * 100, 1) if total > 0 else 0
        result.append({
            "student_id": student.id,
            "student_name": student.full_name,
            "month": month,
            "total_sessions": total,
            "present": present,
            "absent": total - present,
            "percentage": percentage,
        })
    return result


# ── Subject-wise report ───────────────────────────────────────────────────────

@router.get("/subject/{subject_id}")
def subject_report(subject_id: int, db: Session = Depends(get_db)):
    """
    Returns per-student attendance percentage for a specific subject.
    """
    subject = db.query(Subject).filter(Subject.id == subject_id).first()
    if not subject:
        raise HTTPException(status_code=404, detail="Subject not found")

    sessions = db.query(AttendanceSession).filter(
        AttendanceSession.subject_id == subject_id
    ).all()
    session_ids = [s.id for s in sessions]

    students = db.query(User).filter(User.role == "student").all()
    result = []
    for student in students:
        if session_ids:
            total = db.query(AttendanceRecord).filter(
                AttendanceRecord.student_id == student.id,
                AttendanceRecord.session_id.in_(session_ids)
            ).count()
            present = db.query(AttendanceRecord).filter(
                AttendanceRecord.student_id == student.id,
                AttendanceRecord.session_id.in_(session_ids),
                AttendanceRecord.is_present == True
            ).count()
        else:
            total = present = 0

        percentage = round(present / total * 100, 1) if total > 0 else 0
        result.append({
            "student_id": student.id,
            "student_name": student.full_name,
            "subject": subject.name,
            "total_sessions": total,
            "present": present,
            "absent": total - present,
            "percentage": percentage,
        })
    return result


# ── Student-wise report ───────────────────────────────────────────────────────

@router.get("/student/{student_id}")
def student_report(student_id: int, db: Session = Depends(get_db)):
    """
    Returns per-subject attendance breakdown for a specific student.
    """
    student = db.query(User).filter(User.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    subjects = db.query(Subject).filter(Subject.is_active == True).all()
    result = []
    for subject in subjects:
        sessions = db.query(AttendanceSession).filter(
            AttendanceSession.subject_id == subject.id
        ).all()
        session_ids = [s.id for s in sessions]

        if session_ids:
            total = db.query(AttendanceRecord).filter(
                AttendanceRecord.student_id == student_id,
                AttendanceRecord.session_id.in_(session_ids)
            ).count()
            present = db.query(AttendanceRecord).filter(
                AttendanceRecord.student_id == student_id,
                AttendanceRecord.session_id.in_(session_ids),
                AttendanceRecord.is_present == True
            ).count()
        else:
            total = present = 0

        percentage = round(present / total * 100, 1) if total > 0 else 0
        result.append({
            "subject_id": subject.id,
            "subject_name": subject.name,
            "subject_code": subject.code,
            "total_sessions": total,
            "present": present,
            "absent": total - present,
            "percentage": percentage,
        })
    return result


# ── Excel export ──────────────────────────────────────────────────────────────

@router.get("/excel")
def download_excel(db: Session = Depends(get_db)):
    data = get_attendance_data(db)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    headers = ["Name", "Email", "Total Sessions", "Present", "Absent", "Percentage"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for row, student in enumerate(data, 2):
        ws.cell(row=row, column=1, value=student["name"])
        ws.cell(row=row, column=2, value=student["email"])
        ws.cell(row=row, column=3, value=student["total"])
        ws.cell(row=row, column=4, value=student["present"])
        ws.cell(row=row, column=5, value=student["absent"])
        ws.cell(row=row, column=6, value=f"{student['percentage']}%")

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=attendance_report.xlsx"}
    )


# ── PDF export ────────────────────────────────────────────────────────────────

@router.get("/pdf")
def download_pdf(db: Session = Depends(get_db)):
    data = get_attendance_data(db)
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = [Paragraph("Attendance Report", styles["Title"])]

    table_data = [["Name", "Email", "Total", "Present", "Absent", "%"]]
    for student in data:
        table_data.append([
            student["name"], student["email"],
            str(student["total"]), str(student["present"]),
            str(student["absent"]), f"{student['percentage']}%"
        ])

    table = Table(table_data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 12),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF2FF")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("PADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(table)
    doc.build(elements)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=attendance_report.pdf"}
    )
